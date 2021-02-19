# !/usr/bin/env python
# encoding: utf-8
"""
@Time : 2021-01-14 下午 14:17
@project : api_frame
@Author  : xhb
@Site    : 
@File    : logger.py
@Software: PyCharm Community Edition
"""

from openpyxl import load_workbook
import os

file = os.path.dirname(os.getcwd()) + r"\test_data\test_data.xlsx"


class DoExcelOne:
    def __init__(self, sheet):
        # self.file = file
        self.sheet = sheet

    # 获取Excel文件的头部信息
    def get_header(self):
        wb = load_workbook(file)
        sheet = wb[self.sheet]
        header = []
        for i in range(1, sheet.max_column+1):
            header.append(sheet.cell(1, i).value)
        return header

    def get_data(self):
        pass


if __name__ == '__main__':
    print(DoExcelOne("login").get_header())