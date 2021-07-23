# !/usr/bin/env python
# encoding: utf-8
"""
@Time : 2021-04-01 上午 10:10
@project : api_frame
@Author  : xhb
@Site    : 
@File    : aa.py
@Software: PyCharm Community Edition
"""

from openpyxl import load_workbook
import os
from common.project_path import *

filename = BASH_PATH + r"\test_data" + r"\test_data.xlsx"
print(filename)

wb = load_workbook(filename)
sheets = wb.sheetnames
for sheet in sheets:
    print(sheet)