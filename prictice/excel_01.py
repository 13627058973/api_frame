# !/usr/bin/env python
# encoding: utf-8
"""
@Time : 2021-01-08 上午 10:56
@project : api_frame
@Author  : xhb
@Site    : 
@File    : excel_01.py
@Software: PyCharm Community Edition
"""
import item as item
from openpyxl import load_workbook
"""打开文件编辑类"""
"""
# 打开已有的文件
wb = load_workbook("../test_data/test_data.xlsx")
# 打开文件的工作簿（表）
sheet = wb["login"]

# 获取文件的值
print(sheet.cell(2, 1).value)

# 获取最大行，最大列
print(sheet.max_row, sheet.max_column)
"""
"""读取 文件的所有内容"""

# 方法一
wb = load_workbook("../test_data/test_data.xlsx")
sheet = wb["login"]
data = []
for i in range(2, sheet.max_column):
    item = {}
    item["url"] = sheet.cell(i, 1).value
    item["data"] = sheet.cell(i, 2).value
    item["title"] = sheet.cell(i, 3).value
    item["method"] = sheet.cell(i, 4).value
    item["except"] = sheet.cell(i, 5).value
    item["result"] = sheet.cell(i, 6).value
    data.append(item)
print(data)
# 写入内容
sheet.cell(2, 6).value = "111"
wb.save("../test_data/test_data.xlsx")
print(data)
