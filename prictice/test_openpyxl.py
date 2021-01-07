# !/usr/bin/env python
# encoding: utf-8
"""
@Time : 2021-01-07 下午 15:22
@project : api_frame
@Author  : xhb
@Site    : 
@File    : test_openpyxl.py
@Software: PyCharm Community Edition
"""

from openpyxl import Workbook  # 创建文件
from openpyxl import load_workbook  # 打开已有的文件

"""创建类"""
"""
# 创建文件
work = Workbook()
res = work.save("aa.xlsx")
# 创建表单
ws1 = work.create_sheet("python", 0)
# 复制表单
ws = work.copy_worksheet(ws1)
work.save("bb.xlsx")
"""


"""打开文件编辑类"""

# 打开已有的文件
wb = load_workbook("../test_data/test_data.xlsx")
# 打开文件的工作簿（表）
sheet = wb["login"]

# 获取文件的值
print(sheet.cell(2, 1).value)

# 获取最大行，最大列
print(sheet.max_row, sheet.max_column)
