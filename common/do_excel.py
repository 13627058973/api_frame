# !/usr/bin/env python
# encoding: utf-8
"""
@Time : 2021-01-07 下午 15:14
@project : api_frame
@Author  : xhb
@Site    : 
@File    : do_excel.py
@Software: PyCharm Community Edition
"""

from openpyxl import load_workbook
from common.read_conf import *
from common.project_path import *


class DoExcel:

    def __init__(self, file):
        self.file = file

    def get_data(self):  # 获取表单的数据
        wb = load_workbook(self.file)
        mode = eval(ReadConfig.read_config(CONF_PATH, "MODE", "mode"))
        data = []
        tel = int(load_workbook(self.file)["init"].cell(2, 1).value)  # 获取单元格的手机号
        for key in mode:  # 根据配置文件获取需要测试的sheet
            sheet = wb[key]  # 获取表单名
            if mode[key] == "all":
                for i in range(2, sheet.max_row + 1):
                    item = {}
                    item["case_id"] = sheet.cell(i, 1).value
                    item["url"] = sheet.cell(i, 2).value
                    if sheet.cell(i, 3).value.find("${id}") != -1:  # 如果有找到这个${id}
                        item["data"] = sheet.cell(i, 3).value.replace("${id}", tel)
                    else:  # 如果没找到${id}
                        item["data"] = sheet.cell(i, 3).value
                    item["title"] = sheet.cell(i, 4).value
                    item["method"] = sheet.cell(i, 5).value
                    item["type"] = sheet.cell(i, 6).value
                    item["ExpectResult"] = sheet.cell(i, 7).value
                    item["sheet_name"] = key
                    data.append(item)
                    self.update_tel(str(tel+1), self.file, "init")

            else:
                for case_id in mode[key]:
                    item = {}
                    item["case_id"] = sheet.cell(case_id+1, 1).value
                    item["url"] = sheet.cell(case_id+1, 2).value
                    if sheet.cell(case_id+1, 3).value.find("${id}") != -1:  # 如果有找到这个${id}
                        item["data"] = sheet.cell(case_id+1, 3).value.replace("${id}", str(tel))  # 找到了就做数据替换
                    else:  # 如果没找到${id}
                        item["data"] = sheet.cell(i, 3).value
                    item["title"] = sheet.cell(case_id+1, 4).value
                    item["method"] = sheet.cell(case_id+1, 5).value
                    item["type"] = sheet.cell(case_id+1, 6).value
                    item["ExpectResult"] = sheet.cell(case_id+1, 7).value
                    item["sheet_name"] = key
                    data.append(item)
                    self.update_tel(str(tel+1), self.file, "init")  # 更新手机号
        return data

    def write_back(self, sheet_name, row, columns, value):  # 写入文件
        wb = load_workbook(self.file)
        sheet = wb[sheet_name]

        # 在表格中赋值
        sheet.cell(row, columns).value = value

        # 保存文件
        wb.save(self.file)

    @staticmethod
    def update_tel(tel, filename, sheet_name):
        # 更新EXCEL的手机号
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(filename)


if __name__ == '__main__':
    test_data = DoExcel(DATA_PATH).get_data()
    for item in test_data:
        print(item)

